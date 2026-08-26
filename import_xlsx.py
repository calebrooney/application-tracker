"""One-off importer for the existing tracking spreadsheet.

Reads `Sheet1` of Intern Apps.xlsx and inserts each row into Postgres so
existing applications carry over. Run once: `python import_xlsx.py`.

Column mapping (Sheet1):
  A role | B company | C energy? (0/1) | D location | E location type
  F pay | G app date | H job id | J due date | K link
"""

import datetime

from openpyxl import load_workbook

import db

WORKBOOK = "Intern Apps.xlsx"
SHEET = "Sheet1"


def to_date(value):
    """Normalize a cell into a date string (YYYY-MM-DD) or None.

    openpyxl usually returns datetime objects for date-formatted cells,
    but falls back to Excel serial numbers if a cell is plain numeric.
    """
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        # Excel serial dates count days from 1899-12-30.
        base = datetime.date(1899, 12, 30)
        return (base + datetime.timedelta(days=int(value))).isoformat()
    return None


def to_text(value):
    """Return a stripped string, or None for empty cells."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def main():
    """Load Sheet1 rows into the applications table."""
    db.init_db()
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb[SHEET]

    count = 0
    # min_row=2 skips the header row.
    for row in ws.iter_rows(min_row=2, max_col=11, values_only=True):
        role, company, energy, location, loc_type, pay, app_date, job_id, _i, due_date, link = row

        # Skip fully empty rows.
        if not any([role, company, location, link]):
            continue

        data = {
            "role": to_text(role),
            "company": to_text(company),
            "location": to_text(location),
            "location_type": to_text(loc_type),
            "pay": to_text(pay),
            "app_date": to_date(app_date),
            "due_date": to_date(due_date),
            "job_id": to_text(job_id),
            "link": to_text(link),
            "energy_related": str(energy).strip() in ("1", "1.0"),
            "status": "applied",
        }
        db.add_application(data)
        count += 1

    print(f"Imported {count} applications from {WORKBOOK}.")


if __name__ == "__main__":
    main()
